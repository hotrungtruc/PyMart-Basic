from customtkinter import *
from PIL import Image, ImageTk
import openpyxl
from tkinter import messagebox as ctkmessagebox
import subprocess
from tkinter import filedialog
import os

root = CTk()
root.geometry("1200x600")
root.resizable(0,0)
root.title("ĐĂNG NHẬP")

side_img_data = Image.open(r"hinh_anh/dang_nhap/1020353.jpg")
side_img = CTkImage(dark_image=side_img_data, light_image=side_img_data, size=(1200, 800))

def close_frame(frame):
    frame.destroy()


def upload_image():
    global image_file_path, upload_image_button, photo_image
    file_dialog = filedialog.askopenfilename(title="Select Image File", filetypes=[("Image Files", "*.jpg *.jpeg *.png")])
    if file_dialog:
        image_file_path = file_dialog
        image = Image.open(file_dialog)
        image = image.resize((150, 200))
        photo_image = ImageTk.PhotoImage(image)
        upload_image_button.configure(image=photo_image, bg_color="transparent", text = None)
        upload_image_button.image = photo_image  
        upload_image_button.update()

        # Download the image to the desired folder
        folder_path = "hinh_anh/dang_nhap/users"
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        file_name = os.path.basename(file_dialog)
        new_file_path = os.path.join(folder_path, file_name)
        image.save(new_file_path)

upload_image_button = None

def taoTaiKhoan():
    global upload_image_button
    frameTao= CTkFrame(master=root, width=620, height=450, fg_color="#000022")
    frameTao.pack_propagate(0)
    frameTao.place(relx=0.5, rely=0.5, anchor="center")
    CTkLabel(master=frameTao, text="Tạo tài khoản mới", text_color="#1C3AFF", anchor="w", justify="left", font=("Arial Bold", 46)).pack(anchor="center", pady=(25, 0))
    close_button= CTkButton(frameTao, text="X", fg_color="#000000", width =30, height =30,command=lambda: close_frame(frameTao))
    close_button.place(relx=1.0, rely=0.028, anchor="e")
    ten = CTkEntry(master=frameTao, placeholder_text="Tên tài khoản",width=400,height =40, fg_color="#EEEEEE", border_color="#4682B4", border_width=1, text_color="#000000")
    ten.pack(anchor="w", padx=(25, 0), pady=(25,0))
    ID = CTkEntry(master=frameTao, placeholder_text="ID",width=400,height =40, fg_color="#EEEEEE", border_color="#4682B4", border_width=1, text_color="#000000")
    ID.pack(anchor="w", padx=(25, 0), pady=(25,0))
    matKhau = CTkEntry(master=frameTao,placeholder_text="Mật khẩu", width=400,height =40, fg_color="#EEEEEE", border_color="#4682B4", border_width=1, text_color="#000000", show="*")
    matKhau.pack(anchor="w", padx=(25, 0), pady=(25,0))
    nhapLaiMatKhau = CTkEntry(master=frameTao,placeholder_text="Nhập lại mật khẩu", width=400,height =40, fg_color="#EEEEEE", border_color="#4682B4", border_width=1, text_color="#000000", show="*")
    nhapLaiMatKhau.pack(anchor="w", padx=(25, 0), pady=(25,0))
    upload_image_button = CTkButton(master=frameTao, text="Tải ảnh lên", fg_color="#1C3AFF", hover_color="#00008B", font=("Arial Bold", 16), text_color="#ffffff", width=150,height = 200, command = upload_image, corner_radius= 10 ,bg_color = "transparent")
    upload_image_button.place(relx=0.84, rely=0.7, anchor="s")
    CTkButton(master=frameTao, text="Tạo tài khoản", fg_color="#1C3AFF", hover_color="#00008B", font=("Arial Bold", 16), text_color="#ffffff", width=225, command=lambda: save_account(ID.get(), ten.get(), matKhau.get(), nhapLaiMatKhau.get(), image_file_path)).pack(anchor="nw", pady=(40, 0), padx=(25, 0))

def save_account(ID, ten, matKhau, nhapLaiMatKhau, image_file_path):
    if ten == "" or matKhau=="" or ID=="":
        ctkmessagebox.showerror("Error", "Vui lòng điền đầy đủ thông tin")
        return
    if matKhau != nhapLaiMatKhau:
        ctkmessagebox.showwarning("Warning", "The passwords do not match. Please try again.")
        return
    try:
        workbook = openpyxl.load_workbook("dang_nhap\Logins.xlsx")
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
    worksheet = workbook.active
    max_row = worksheet.max_row
    for row in range(1, max_row + 1):
        if worksheet.cell(row=row, column=2).value == ten:
            ctkmessagebox.showwarning("Warning", "The account name already exists. Please choose a different account name.")
            return
        
    for row in range(1, max_row + 1):
        if worksheet.cell(row=row, column=1).value == str(ID):
            ctkmessagebox.showwarning("Warning", "The account ID already exists. Please choose a different account ID.")
            return
        
    worksheet.cell(row=max_row + 1, column=1).value = ID  # save as "id"
    worksheet.cell(row=max_row + 1, column=2).value = image_file_path  # save as "image"
    worksheet.cell(row=max_row + 1, column=3).value = ten  # save as "name"
    worksheet.cell(row=max_row + 1, column=4).value = matKhau  # save as "password"
    workbook.save("dang_nhap\Logins.xlsx")
    ctkmessagebox.showinfo("Success", "Registered successfully!")
    
def check_login(account, password):
    try:
        workbook = openpyxl.load_workbook("dang_nhap\Logins.xlsx")
    except FileNotFoundError:
        return False

    worksheet = workbook.active
    max_row = worksheet.max_row

    for row in range(1, max_row + 1):
        if worksheet.cell(row=row, column=3).value == account and worksheet.cell(row=row, column=4).value == password:
            return True
    return False


def login():
    account = ten.get()
    password = matKhau.get()

    if check_login(account, password):
        # Run another Python file
        subprocess.Popen(["python", "view.py"])
        root.destroy()
    else:
        # Show a warning window
        ctkmessagebox.showwarning("Warning", "Invalid account or password")
CTkLabel(master=root, text="", image=side_img).place(relx=0.5, rely=0.5, anchor="center")


frame = CTkFrame(master=root, width=500, height=400, fg_color="#000022")
frame.pack_propagate(0)
frame.place(relx=0.5, rely=0.5, anchor="center")

CTkLabel(master=frame, text="Welcome Back!", text_color="#1C3AFF", anchor="w", justify="left", font=("Arial Bold", 46)).pack(anchor="center", pady=(25, 0))
CTkLabel(master=frame, text="Đăng nhập vào tài khoản của bạn.", text_color="#00CDCD", anchor="w", justify="left", font=("Arial Bold",16)).pack(anchor="center")

ten = CTkEntry(master=frame, placeholder_text="Tên tài khoản",width=400, height =40, fg_color="#EEEEEE", border_color="#4682B4", border_width=1, text_color="#000000")
ten.pack(anchor="center",pady=(35,0))

matKhau = CTkEntry(master=frame,placeholder_text="Mật khẩu", width=400, height = 40, fg_color="#EEEEEE", border_color="#4682B4", border_width=1, text_color="#000000", show="*")
matKhau.pack(anchor="center",pady=(5,0))

CTkButton(master=frame, text="Đăng nhập", fg_color="#1C3AFF", hover_color="#00008B", font=("Arial Bold", 20), text_color="#ffffff", width=400, height=40, command = login).pack(anchor="center", pady=(35, 0))

CTkButton(master=frame, text="Tạo tài khoản", fg_color="#1C3AFF", hover_color="#00008B", font=("Arial Bold", 20), text_color="#ffffff", command=taoTaiKhoan, width=200, height=40).pack(anchor="center", pady=(20, 0))

root.mainloop()