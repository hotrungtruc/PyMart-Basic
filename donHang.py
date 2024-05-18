from tkinter import *
from customtkinter import *
import tkinter.messagebox as messagebox
import random
import tempfile
from openpyxl import Workbook
import pandas as pd
import openpyxl


def close_frame(frame):
    frame.destroy()


qty_product = []

def total_product(aaaEntry, aaEntry, usbcEntry, whEntry, lccEntry, BshEntry, m20Entry, aahEntry, tvEntry, m34Entry, gmEntry, vareebaddphoneEntry, googlephoneEntry, lgdryerEntry, wmEntry, iphoneEntry, thinkpadEntry, macEntry):
    global current_qty_product, total
    current_qty_product = (
        int(aaaEntry.get()) +
        int(aaEntry.get()) +
        int(usbcEntry.get()) +
        int(whEntry.get()) +
        int(lccEntry.get()) +
        int(BshEntry.get()) +
        int(m20Entry.get()) +
        int(aahEntry.get()) +
        int(tvEntry.get()) +
        int(m34Entry.get()) +
        int(gmEntry.get()) +
        int(vareebaddphoneEntry.get()) +
        int(googlephoneEntry.get()) +
        int(lgdryerEntry.get()) +
        int(wmEntry.get()) +
        int(iphoneEntry.get()) +
        int(thinkpadEntry.get()) +
        int(macEntry.get()))
    qty_product.append(current_qty_product)
    total = sum(qty_product)
    return total

def get_total_product():
    return sum(qty_product)







current_sales=0
def total(aaaEntry, aaEntry, usbcEntry, whEntry, lccEntry, BshEntry, m20Entry, aahEntry, tvEntry, m34Entry, gmEntry, vareebaddphoneEntry, googlephoneEntry, lgdryerEntry, wmEntry, iphoneEntry, thinkpadEntry, macEntry, totalEntry,taxentry):
    global aaaprice
    global aaprice
    global usbprice
    global whprice
    global lccprice
    global bshprice
    global m20price
    global aahprice
    global tvprice
    global m34price
    global gmprice
    global vareebaddphoneprice
    global googlephoneprice
    global lgdryerprice
    global wmprice
    global iphoneprice
    global thinkpadprice
    global macprice
    global total_price
    global totalbill
    global current_sales
    aaaprice = int(aaaEntry.get())*2.99
    aaprice = int(aaEntry.get())*3.84
    usbprice = int(usbcEntry.get())*11.95
    whprice = int(whEntry.get())*11.99
    lccprice = int(lccEntry.get())*14.95
    bshprice = int(BshEntry.get())*99.99
    m20price = int(m20Entry.get())*109.99
    aahprice = int(aahEntry.get())*150
    tvprice = int(tvEntry.get())*300
    m34price = int(m34Entry.get())*379.99
    gmprice = int(gmEntry.get())*389.99
    vareebaddphoneprice = int(vareebaddphoneEntry.get())*400
    googlephoneprice = int(googlephoneEntry.get())*600
    lgdryerprice = int(lgdryerEntry.get())*600
    wmprice = int(wmEntry.get())*600
    iphoneprice = int(iphoneEntry.get())*700
    thinkpadprice = int(thinkpadEntry.get())*999.99
    macprice = int(macEntry.get())*1700
    total_price = (
    aaaprice + aaprice + usbprice + whprice + lccprice + bshprice +
    m20price + aahprice + tvprice + m34price + gmprice +
    vareebaddphoneprice + googlephoneprice + lgdryerprice +
    wmprice + iphoneprice + thinkpadprice + macprice)
    totalEntry.delete(0, END)
    totalEntry.insert(0,f'{total_price} $')
    tax(aaaEntry, aaEntry, usbcEntry, whEntry, lccEntry, BshEntry, m20Entry, aahEntry, tvEntry, m34Entry, gmEntry, vareebaddphoneEntry, googlephoneEntry, lgdryerEntry, wmEntry, iphoneEntry, thinkpadEntry, macEntry,taxentry)
    totalbill = total_price - total_tax
    current_sales=current_sales+total_price
    total_product(aaaEntry, aaEntry, usbcEntry, whEntry, lccEntry, BshEntry, m20Entry, aahEntry, tvEntry, m34Entry, gmEntry, vareebaddphoneEntry, googlephoneEntry, lgdryerEntry, wmEntry, iphoneEntry, thinkpadEntry, macEntry)

def get_current_sales():
    return current_sales

def tax(aaaEntry, aaEntry, usbcEntry, whEntry, lccEntry, BshEntry, m20Entry, aahEntry, tvEntry, m34Entry, gmEntry, vareebaddphoneEntry, googlephoneEntry, lgdryerEntry, wmEntry, iphoneEntry, thinkpadEntry, macEntry, taxentry):
    global aaatax
    global aatax
    global usbtax
    global whtax
    global lcctax
    global bshtax
    global m20tax
    global aahtax
    global tvtax
    global m34tax
    global gmtax
    global vareebaddphonetax
    global googlephonetax
    global lgdryertax
    global wmtax
    global iphonetax
    global thinkpadtax
    global mactax
    global total_tax
    aaatax = int(aaaEntry.get())*2.99*0.1
    aatax = int(aaEntry.get())*3.84*0.1
    usbtax = int(usbcEntry.get())*11.95*0.1
    whtax = int(whEntry.get())*11.99*0.1
    lcctax = int(lccEntry.get())*14.95*0.1
    bshtax = int(BshEntry.get())*99.99*0.1
    m20tax = int(m20Entry.get())*109.99*0.1
    aahtax = int(aahEntry.get())*150*0.1
    tvtax = int(tvEntry.get())*300*0.1
    m34tax = int(m34Entry.get())*379.99*0.1
    gmtax = int(gmEntry.get())*389.99*0.1
    vareebaddphonetax = int(vareebaddphoneEntry.get())*400*0.1
    googlephonetax = int(googlephoneEntry.get())*600*0.1
    lgdryertax = int(lgdryerEntry.get())*600*0.1
    wmtax = int(wmEntry.get())*600*0.1
    iphonetax = int(iphoneEntry.get())*700*0.1
    thinkpadtax = int(thinkpadEntry.get())*999.99*0.1
    mactax = int(macEntry.get())*1700*0.1
    total_tax = (
    aaatax + aatax + usbtax + whtax + lcctax + bshtax +
    m20tax + aahtax + tvtax + m34tax + gmtax +
    vareebaddphonetax + googlephonetax + lgdryertax +
    wmtax + iphonetax + thinkpadtax + mactax)
    taxentry.delete(0, END)
    taxentry.insert(0,f'{total_tax} $')


def bill_area(nameEntry, totalEntry, phoneEntry, textarea, aaaEntry, aaEntry, usbcEntry, whEntry, lccEntry, BshEntry, m20Entry, aahEntry, tvEntry, m34Entry, gmEntry, vareebaddphoneEntry, googlephoneEntry, lgdryerEntry, wmEntry, iphoneEntry, thinkpadEntry, macEntry,taxentry):
        if nameEntry.get() == '' or phoneEntry.get() == '':
            messagebox.showerror('Error', "Customer details are required!!!")
        elif totalEntry.get() == '' or totalEntry.get() == '0.0 $':
            messagebox.showerror('Error', "No product selected")
        else:
            textarea.insert(END, '\t\t**Welcome customer**')
            textarea.insert(END, f'\nBill Number: {billnumber}\n')
            textarea.insert(END, f'\nCustomer Name: {nameEntry.get()}\n')
            textarea.insert(END, f'\nCustomer Phone Number: {phoneEntry.get()}\n')
            textarea.insert(END,'\n======================================================')
            textarea.insert(END, '\tProduct\t\t\tQuantity\t\tPrice')
            textarea.insert(END,'\n======================================================')
            if int(aaaEntry.get()) != 0:
                textarea.insert(END, f'\nAAA Batteries (4-pack)\t\t\t{aaaEntry.get()}\t\t{aaaprice}')
            if int(aaEntry.get()) != 0:
                textarea.insert(END, f'\nAA Batteries (4-pack)\t\t\t{aaEntry.get()}\t\t{aaprice}')
            if int(usbcEntry.get()) != 0:
                textarea.insert(END, f'\nUSB-C Charging Cable\t\t\t{usbcEntry.get()}\t\t{usbprice}')
            if int(whEntry.get()) != 0:
                textarea.insert(END, f'\nWired Headphones\t\t\t{whEntry.get()}\t\t{whprice}')
            if int(lccEntry.get()) != 0:
                textarea.insert(END, f'\nLightning Charging Cable\t\t\t{lccEntry.get()}\t\t{lccprice}')
            if int(BshEntry.get()) != 0:
                textarea.insert(END, f'\nBose SoundSport Headphones\t\t\t{BshEntry.get()}\t\t{bshprice}')
            if int(m20Entry.get()) != 0:
                textarea.insert(END, f'\n20in Monitor\t\t\t{m20Entry.get()}\t\t{m20price}')
            if int(aahEntry.get()) != 0:
                textarea.insert(END, f'\n27in FHD Monitor\t\t\t{aahEntry.get()}\t\t{aahprice}')
            if int(tvEntry.get()) != 0:
                textarea.insert(END, f'\nFlatscreen TV\t\t\t{tvEntry.get()}\t\t{tvprice}')
            if int(m34Entry.get()) != 0:
                textarea.insert(END, f'\n34in Ultrawide Monitor\t\t\t{m34Entry.get()}\t\t{m34price}')
            if int(gmEntry.get()) != 0:
                textarea.insert(END, f'\n27in 4K Gaming Monitor\t\t\t{gmEntry.get()}\t\t{gmprice}')
            if int(vareebaddphoneEntry.get()) != 0:
                textarea.insert(END, f'\nVareebadd Phone\t\t\t{vareebaddphoneEntry.get()}\t\t{vareebaddphoneprice}')
            if int(googlephoneEntry.get()) != 0:
                textarea.insert(END, f'\nGoogle Phone\t\t\t{googlephoneEntry.get()}\t\t{googlephoneprice}')
            if int(lgdryerEntry.get()) != 0:
                textarea.insert(END, f'\nLG Dryer\t\t\t{lgdryerEntry.get()}\t\t{lgdryerprice}')
            if int(wmEntry.get()) != 0:
                textarea.insert(END, f'\nLG Washing Machine\t\t\t{wmEntry.get()}\t\t{wmprice}')
            if int(iphoneEntry.get()) != 0:
                textarea.insert(END, f'\nIphone\t\t\t{iphoneEntry.get()}\t\t{iphoneprice}')
            if int(thinkpadEntry.get()) != 0:
                textarea.insert(END, f'\nThinkPad Laptops\t\t\t{thinkpadEntry.get()}\t\t{thinkpadprice}')
            if int(macEntry.get()) != 0:
                textarea.insert(END, f'\nMacbook Pro Laptops\t\t\t{macEntry.get()}\t\t{macprice}')
            textarea.insert(END,'\n-------------------------------------------------------')
            textarea.insert(END, f'\nTAX\t\t\t{taxentry.get()}')
            textarea.insert(END, f'\nTOTAL\t\t\t{totalbill}')
            save_bill(textarea)
            save_data_customer(nameEntry, phoneEntry,totalbill)



if not os.path.exists('bills'):
    os.mkdir('bills')

billno=0
def save_bill(textarea) :
    global billnumber, billno
    result = messagebox.askyesno('Cofirm', "Do you want to save bill?")
    if result:
        bill_content = textarea.get(1.0, END)
        filename = f'bills/{billnumber}.txt'
        with open(filename, 'w') as file:
            file.write(bill_content)
        messagebox.showinfo('Success', f'{billnumber} is saved successfully')
        billnumber = random.randint(1000, 9999)
        billno= billno+1

billnumber = random.randint(1000, 9999)

def get_bill_count():
    return billno

def search_bill(billnumberEntry, textarea):
    for i in os.listdir('DATABASE\bills'):
        if i.split('.')[0]==billnumberEntry.get():
            with open(f'bills/{i}', 'r') as f:
                textarea.delete(1.0, END)
                for data in f:
                    textarea.insert(END,data)
                f.close()
                break
    else:
        messagebox.showerror('Error','Invalid Bill Number')



def print_bill(textarea):
    if textarea.get(1.0, END)=='\n':
        messagebox.showerror('Error', "Bill is empty")
    else:
        file=tempfile.mktemp('.txt')
        with open(file,'w') as f:
            f.write(textarea.get(1.0, END))
        os.startfile(file, 'print')



def save_data_customer(nameEntry, phoneEntry, totalbill):
    try:
        workbook = openpyxl.load_workbook(r"Database\Customer.xlsx")
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
    worksheet = workbook.active
    max_row = worksheet.max_row
    worksheet.cell(row=max_row + 1, column=1).value = billnumber
    worksheet.cell(row=max_row + 1, column=2).value = nameEntry.get()
    worksheet.cell(row=max_row + 1, column=3).value = phoneEntry.get()
    worksheet.cell(row=max_row + 1, column=4).value = totalbill
    workbook.save(r"Database\Customer.xlsx")
   